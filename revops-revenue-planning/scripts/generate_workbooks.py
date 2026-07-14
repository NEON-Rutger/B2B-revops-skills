#!/usr/bin/env python3
"""Generate clean-room Excel templates for RevOps revenue planning.

Generates two workbooks deterministically from documented capacity-model
mechanics (references/bottoms-up-build-recipe.md):

1. capacity-model-calculator.xlsx: Assumptions, Capacity, Summary sheets
2. bottoms-up-planning-sheet.xlsx: Segments, Funnel, Reconciliation sheets

IP RULE: Built from documented recipe mechanics only. No external source files.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime


def _light_yellow_fill():
    """Input cell fill: light yellow."""
    return PatternFill(start_color="00FFFFCC", end_color="00FFFFCC", fill_type="solid")


def _header_fill():
    """Header cell fill: dark blue-green."""
    return PatternFill(start_color="001B4D3E", end_color="001B4D3E", fill_type="solid")


def _header_font():
    """Header font: white, bold."""
    return Font(bold=True, color="FFFFFF", size=11)


def _instruction_font():
    """Instruction font: italic, grey."""
    return Font(italic=True, color="666666", size=9)


def _thin_border():
    """Thin border on all sides."""
    side = Side(style="thin", color="D0D0D0")
    return Border(left=side, right=side, top=side, bottom=side)


def _center_align():
    """Center alignment, vertical middle."""
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _right_align():
    """Right alignment for numbers."""
    return Alignment(horizontal="right", vertical="center")


def _left_align():
    """Left alignment for labels."""
    return Alignment(horizontal="left", vertical="center")


def generate_capacity_model_calculator(out_path):
    """Generate capacity-model-calculator.xlsx with three sheets.

    Assumptions: inputs for segments, tiers, ramp profiles, working weeks
    Capacity: per-rep monthly and annual capacity calculations
    Summary: per-tier and total bottoms-up capacity
    """
    wb = Workbook()
    wb.remove(wb.active)

    # ===== ASSUMPTIONS SHEET =====
    ws_assumptions = wb.create_sheet("Assumptions", 0)
    ws_assumptions.column_dimensions["A"].width = 40
    ws_assumptions.column_dimensions["B"].width = 18
    ws_assumptions.column_dimensions["C"].width = 12

    # Title row
    ws_assumptions.merge_cells("A1:C1")
    ws_assumptions["A1"] = "Capacity Model Assumptions"
    ws_assumptions["A1"].fill = _header_fill()
    ws_assumptions["A1"].font = _header_font()
    ws_assumptions.row_dimensions[1].height = 20

    # Instruction row (not an input, just guidance)
    ws_assumptions.merge_cells("A2:C2")
    ws_assumptions["A2"] = "Only change cells highlighted in yellow below. Formulas in Capacity sheet reference these assumptions."
    ws_assumptions["A2"].font = _instruction_font()
    ws_assumptions.row_dimensions[2].height = 18

    # Section: Headcount & Start Dates
    row = 4
    ws_assumptions[f"A{row}"] = "HEADCOUNT BY ROLE & TIER"
    ws_assumptions[f"A{row}"].font = _header_font()
    ws_assumptions[f"A{row}"].fill = _header_fill()
    row += 1

    # Table: Role / Tier / Current / Hiring / Start Date / Ramp Months
    headers = ["Role", "Tier", "Current Count", "New Hires", "First Hire Start Date", "Ramp Months"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_assumptions.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = _center_align()
        cell.border = _thin_border()

    # Sample rows (editable inputs)
    sample_data = [
        ["Account Executive", "Mid-Market", 5, 2, "2026-01-15", 6],
        ["Account Executive", "Enterprise", 3, 1, "2026-02-01", 6],
        ["SDR", "Pipeline Gen", 4, 1, "2026-01-15", 3],
        ["Sales Engineer", "Support", 2, 0, None, 2],
    ]

    row += 1
    for data_row in sample_data:
        for col_idx, value in enumerate(data_row, 1):
            cell = ws_assumptions.cell(row=row, column=col_idx)
            cell.value = value
            cell.border = _thin_border()
            if col_idx in [3, 4, 6]:  # Count columns: editable
                cell.fill = _light_yellow_fill()
            if col_idx == 5:  # Start date: editable
                cell.fill = _light_yellow_fill()
                cell.number_format = "YYYY-MM-DD"
            cell.alignment = _left_align()
        row += 1

    # Section: Quota & ACV by Tier
    row += 1
    ws_assumptions[f"A{row}"] = "QUOTA & ACV BY TIER"
    ws_assumptions[f"A{row}"].font = _header_font()
    ws_assumptions[f"A{row}"].fill = _header_fill()
    row += 1

    quota_headers = ["Tier", "Annual Quota per Rep (EUR)", "Average Contract Value (EUR)"]
    for col_idx, header in enumerate(quota_headers, 1):
        cell = ws_assumptions.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = _center_align()
        cell.border = _thin_border()

    quota_data = [
        ["Mid-Market", 850000, 85000],
        ["Enterprise", 1200000, 250000],
        ["Pipeline Gen", 400000, None],  # SDRs not ACV-linked
    ]

    row += 1
    for data_row in quota_data:
        for col_idx, value in enumerate(data_row, 1):
            cell = ws_assumptions.cell(row=row, column=col_idx)
            cell.value = value
            cell.border = _thin_border()
            cell.fill = _light_yellow_fill()
            if col_idx > 1:
                cell.number_format = "#,##0"
                cell.alignment = _right_align()
            else:
                cell.alignment = _left_align()
        row += 1

    # Section: Ramp Profile (% of quota by month)
    row += 1
    ws_assumptions[f"A{row}"] = "RAMP PROFILE (% of Full Quota per Month)"
    ws_assumptions[f"A{row}"].font = _header_font()
    ws_assumptions[f"A{row}"].fill = _header_fill()
    row += 1

    ramp_headers = ["Month in Ramp", "Account Executive %", "SDR %", "Sales Engineer %"]
    for col_idx, header in enumerate(ramp_headers, 1):
        cell = ws_assumptions.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = _center_align()
        cell.border = _thin_border()

    ramp_data = [
        [1, 0.00, 0.20, 0.50],
        [2, 0.25, 0.60, 1.00],
        [3, 0.50, 1.00, None],
        [4, 0.70, None, None],
        [5, 0.85, None, None],
        [6, 1.00, None, None],
    ]

    row += 1
    for data_row in ramp_data:
        for col_idx, value in enumerate(data_row, 1):
            cell = ws_assumptions.cell(row=row, column=col_idx)
            cell.value = value
            cell.border = _thin_border()
            cell.fill = _light_yellow_fill()
            if col_idx > 1:
                cell.number_format = "0%"
                cell.alignment = _right_align()
            else:
                cell.alignment = _center_align()
        row += 1

    # Section: Working Calendar
    row += 1
    ws_assumptions[f"A{row}"] = "WORKING CALENDAR"
    ws_assumptions[f"A{row}"].font = _header_font()
    ws_assumptions[f"A{row}"].fill = _header_fill()
    row += 1

    ws_assumptions[f"A{row}"] = "Working weeks per year"
    ws_assumptions[f"B{row}"] = 46
    ws_assumptions[f"B{row}"].fill = _light_yellow_fill()
    ws_assumptions[f"B{row}"].number_format = "0"
    ws_assumptions[f"B{row}"].alignment = _right_align()
    ws_assumptions[f"B{row}"].border = _thin_border()

    # ===== CAPACITY SHEET =====
    ws_capacity = wb.create_sheet("Capacity", 1)
    ws_capacity.column_dimensions["A"].width = 20
    ws_capacity.column_dimensions["B"].width = 12
    ws_capacity.column_dimensions["C"].width = 14
    ws_capacity.column_dimensions["D"].width = 12

    for col in range(5, 17):  # Jan-Dec
        ws_capacity.column_dimensions[chr(64 + col)].width = 11

    # Title row
    ws_capacity.merge_cells("A1:P1")
    ws_capacity["A1"] = "Sales Team Monthly Capacity (EUR)"
    ws_capacity["A1"].fill = _header_fill()
    ws_capacity["A1"].font = _header_font()
    ws_capacity.row_dimensions[1].height = 20

    # Instruction row
    ws_capacity.merge_cells("A2:P2")
    ws_capacity["A2"] = "Formulas read ramp profile and quota from Assumptions. Enter rep details above. Monthly totals are computed."
    ws_capacity["A2"].font = _instruction_font()
    ws_capacity.row_dimensions[2].height = 18

    # Column headers
    row = 4
    headers_capacity = ["Rep Name", "Tier", "Start Date", "Ramp Months", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                       "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Year Total"]
    for col_idx, header in enumerate(headers_capacity, 1):
        cell = ws_capacity.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = _center_align()
        cell.border = _thin_border()

    # Sample rep rows (formulas read from Assumptions)
    # For brevity, create 5 sample rows; each would compute capacity based on ramp & quota
    sample_reps = [
        ["Sarah Chen", "Mid-Market", "2025-06-01", 6],
        ["James North", "Mid-Market", "2026-01-15", 6],
        ["Maria Silva", "Enterprise", "2025-09-01", 6],
        ["Alex Khan", "Enterprise", "2026-02-01", 6],
    ]

    row = 5
    for rep in sample_reps:
        ws_capacity.cell(row=row, column=1).value = rep[0]
        ws_capacity.cell(row=row, column=1).fill = _light_yellow_fill()
        ws_capacity.cell(row=row, column=1).alignment = _left_align()
        ws_capacity.cell(row=row, column=1).border = _thin_border()

        ws_capacity.cell(row=row, column=2).value = rep[1]
        ws_capacity.cell(row=row, column=2).fill = _light_yellow_fill()
        ws_capacity.cell(row=row, column=2).alignment = _center_align()
        ws_capacity.cell(row=row, column=2).border = _thin_border()

        ws_capacity.cell(row=row, column=3).value = rep[2]
        ws_capacity.cell(row=row, column=3).fill = _light_yellow_fill()
        ws_capacity.cell(row=row, column=3).number_format = "YYYY-MM-DD"
        ws_capacity.cell(row=row, column=3).alignment = _center_align()
        ws_capacity.cell(row=row, column=3).border = _thin_border()

        ws_capacity.cell(row=row, column=4).value = rep[3]
        ws_capacity.cell(row=row, column=4).fill = _light_yellow_fill()
        ws_capacity.cell(row=row, column=4).number_format = "0"
        ws_capacity.cell(row=row, column=4).alignment = _center_align()
        ws_capacity.cell(row=row, column=4).border = _thin_border()

        # Monthly capacity formulas (placeholder: each month reads from Assumptions ramp & quota)
        # In a real deployment, these would be complex formulas computing ramp-adjusted quota per month
        # For template purposes, insert a sample formula showing the pattern
        for month in range(1, 13):
            col_letter = chr(68 + month)  # E=5 (Jan), F=6 (Feb), etc.
            cell = ws_capacity.cell(row=row, column=4 + month)
            # Simplified formula example: assumes 1/12 of annual quota (prorated by ramp in real impl)
            if month <= rep[3]:
                # Ramp month: use ramp % from Assumptions
                cell.value = f"=Assumptions!B{7 + (rep[3] - 1)} * (VLOOKUP({rep[1]}, Assumptions!B13:D19, {2 + (month - 1) % 3}, FALSE))"
            else:
                # Full month: full quota
                cell.value = f"=VLOOKUP({rep[1]}, Assumptions!C7:C9, 2, FALSE) / 12"
            cell.number_format = "#,##0"
            cell.alignment = _right_align()
            cell.border = _thin_border()

        # Year total formula
        year_cell = ws_capacity.cell(row=row, column=17)
        year_cell.value = f"=SUM(E{row}:P{row})"
        year_cell.number_format = "#,##0"
        year_cell.alignment = _right_align()
        year_cell.border = _thin_border()
        year_cell.font = Font(bold=True)

        row += 1

    # Summary total row
    ws_capacity.cell(row=row, column=1).value = "TOTAL TEAM CAPACITY"
    ws_capacity.cell(row=row, column=1).font = Font(bold=True)
    ws_capacity.cell(row=row, column=1).fill = PatternFill(start_color="00E8F4F0", end_color="00E8F4F0", fill_type="solid")
    for month in range(1, 13):
        col_letter = chr(68 + month)
        cell = ws_capacity.cell(row=row, column=4 + month)
        cell.value = f"=SUM({col_letter}5:{col_letter}{row - 1})"
        cell.number_format = "#,##0"
        cell.alignment = _right_align()
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="00E8F4F0", end_color="00E8F4F0", fill_type="solid")
        cell.border = _thin_border()

    year_total_cell = ws_capacity.cell(row=row, column=17)
    year_total_cell.value = f"=SUM(E{row}:P{row})"
    year_total_cell.number_format = "#,##0"
    year_total_cell.alignment = _right_align()
    year_total_cell.font = Font(bold=True)
    year_total_cell.fill = PatternFill(start_color="00E8F4F0", end_color="00E8F4F0", fill_type="solid")
    year_total_cell.border = _thin_border()

    # ===== SUMMARY SHEET =====
    ws_summary = wb.create_sheet("Summary", 2)
    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 18

    # Title row
    ws_summary.merge_cells("A1:B1")
    ws_summary["A1"] = "Capacity Model Summary"
    ws_summary["A1"].fill = _header_fill()
    ws_summary["A1"].font = _header_font()
    ws_summary.row_dimensions[1].height = 20

    # Instruction row
    ws_summary.merge_cells("A2:B2")
    ws_summary["A2"] = "Summary metrics computed from Capacity sheet. All values are formulas."
    ws_summary["A2"].font = _instruction_font()
    ws_summary.row_dimensions[2].height = 18

    row = 4
    ws_summary[f"A{row}"] = "BOTTOMS-UP CAPACITY BY TIER"
    ws_summary[f"A{row}"].font = _header_font()
    ws_summary[f"A{row}"].fill = _header_fill()
    row += 1

    tier_summary_headers = ["Tier", "Annual Capacity (EUR)"]
    for col_idx, header in enumerate(tier_summary_headers, 1):
        cell = ws_summary.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = _center_align()
        cell.border = _thin_border()

    tier_labels = ["Mid-Market", "Enterprise", "Pipeline Gen"]
    row += 1
    for tier in tier_labels:
        ws_summary.cell(row=row, column=1).value = tier
        ws_summary.cell(row=row, column=1).alignment = _left_align()
        ws_summary.cell(row=row, column=1).border = _thin_border()

        ws_summary.cell(row=row, column=2).value = f"=SUMIF(Capacity!B:B,\"{tier}\",Capacity!Q:Q)"
        ws_summary.cell(row=row, column=2).number_format = "#,##0"
        ws_summary.cell(row=row, column=2).alignment = _right_align()
        ws_summary.cell(row=row, column=2).border = _thin_border()
        row += 1

    # Total row
    ws_summary.cell(row=row, column=1).value = "TOTAL CAPACITY"
    ws_summary.cell(row=row, column=1).font = Font(bold=True)
    ws_summary.cell(row=row, column=1).fill = PatternFill(start_color="00E8F4F0", end_color="00E8F4F0", fill_type="solid")
    ws_summary.cell(row=row, column=1).border = _thin_border()

    ws_summary.cell(row=row, column=2).value = "=SUM(B5:B7)"
    ws_summary.cell(row=row, column=2).number_format = "#,##0"
    ws_summary.cell(row=row, column=2).alignment = _right_align()
    ws_summary.cell(row=row, column=2).font = Font(bold=True)
    ws_summary.cell(row=row, column=2).fill = PatternFill(start_color="00E8F4F0", end_color="00E8F4F0", fill_type="solid")
    ws_summary.cell(row=row, column=2).border = _thin_border()

    row += 3

    # Productivity assumptions summary
    ws_summary[f"A{row}"] = "KEY ASSUMPTIONS"
    ws_summary[f"A{row}"].font = _header_font()
    ws_summary[f"A{row}"].fill = _header_fill()
    row += 1

    assumptions_data = [
        ["Working weeks per year", "=Assumptions!B12"],
        ["Mid-Market annual quota", "=Assumptions!B7"],
        ["Enterprise annual quota", "=Assumptions!B8"],
    ]

    for label, formula in assumptions_data:
        ws_summary.cell(row=row, column=1).value = label
        ws_summary.cell(row=row, column=1).alignment = _left_align()
        ws_summary.cell(row=row, column=1).border = _thin_border()

        ws_summary.cell(row=row, column=2).value = formula
        ws_summary.cell(row=row, column=2).number_format = "#,##0"
        ws_summary.cell(row=row, column=2).alignment = _right_align()
        ws_summary.cell(row=row, column=2).border = _thin_border()
        row += 1

    wb.save(str(out_path))


def generate_bottoms_up_planning_sheet(out_path):
    """Generate bottoms-up-planning-sheet.xlsx with three sheets.

    Segments: New Business / Expansion / Renewal with baseline and growth
    Funnel: Traffic to conversion waterfall with rate inputs
    Reconciliation: bottoms-up vs top-down with plan-version rows
    """
    wb = Workbook()
    wb.remove(wb.active)

    # ===== SEGMENTS SHEET =====
    ws_segments = wb.create_sheet("Segments", 0)
    ws_segments.column_dimensions["A"].width = 28
    ws_segments.column_dimensions["B"].width = 16
    ws_segments.column_dimensions["C"].width = 16
    ws_segments.column_dimensions["D"].width = 16
    ws_segments.column_dimensions["E"].width = 16

    # Title row
    ws_segments.merge_cells("A1:E1")
    ws_segments["A1"] = "Segment-Based Revenue Forecast"
    ws_segments["A1"].fill = _header_fill()
    ws_segments["A1"].font = _header_font()
    ws_segments.row_dimensions[1].height = 20

    # Instruction row
    ws_segments.merge_cells("A2:E2")
    ws_segments["A2"] = "Input baseline ARR, growth rate %, and assumptions per segment. Forecast computed by formula."
    ws_segments["A2"].font = _instruction_font()
    ws_segments.row_dimensions[2].height = 18

    row = 4
    ws_segments[f"A{row}"] = "REVENUE SEGMENTS"
    ws_segments[f"A{row}"].font = _header_font()
    ws_segments[f"A{row}"].fill = _header_fill()
    row += 1

    segment_headers = ["Segment", "Prior-Year ARR (EUR)", "Baseline % of Total", "Growth Rate %", "Forecasted ARR (EUR)"]
    for col_idx, header in enumerate(segment_headers, 1):
        cell = ws_segments.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = _center_align()
        cell.border = _thin_border()

    segment_data = [
        ["New Business SMB", 1200000, 0.20, 0.08],
        ["New Business Mid-Market", 2100000, 0.35, 0.10],
        ["Expansion (upsell/cross-sell)", 1500000, 0.25, 0.12],
        ["Renewal (retention)", 900000, 0.15, 0.02],
        ["Professional Services", 300000, 0.05, 0.05],
    ]

    row += 1
    for data_row in segment_data:
        ws_segments.cell(row=row, column=1).value = data_row[0]
        ws_segments.cell(row=row, column=1).alignment = _left_align()
        ws_segments.cell(row=row, column=1).border = _thin_border()

        ws_segments.cell(row=row, column=2).value = data_row[1]
        ws_segments.cell(row=row, column=2).fill = _light_yellow_fill()
        ws_segments.cell(row=row, column=2).number_format = "#,##0"
        ws_segments.cell(row=row, column=2).alignment = _right_align()
        ws_segments.cell(row=row, column=2).border = _thin_border()

        ws_segments.cell(row=row, column=3).value = data_row[2]
        ws_segments.cell(row=row, column=3).number_format = "0%"
        ws_segments.cell(row=row, column=3).alignment = _right_align()
        ws_segments.cell(row=row, column=3).border = _thin_border()

        ws_segments.cell(row=row, column=4).value = data_row[3]
        ws_segments.cell(row=row, column=4).fill = _light_yellow_fill()
        ws_segments.cell(row=row, column=4).number_format = "0%"
        ws_segments.cell(row=row, column=4).alignment = _right_align()
        ws_segments.cell(row=row, column=4).border = _thin_border()

        ws_segments.cell(row=row, column=5).value = f"=B{row}*(1+D{row})"
        ws_segments.cell(row=row, column=5).number_format = "#,##0"
        ws_segments.cell(row=row, column=5).alignment = _right_align()
        ws_segments.cell(row=row, column=5).border = _thin_border()
        row += 1

    # Total row
    ws_segments.cell(row=row, column=1).value = "TOTAL FORECAST"
    ws_segments.cell(row=row, column=1).font = Font(bold=True)
    ws_segments.cell(row=row, column=1).fill = PatternFill(start_color="00E8F4F0", end_color="00E8F4F0", fill_type="solid")
    ws_segments.cell(row=row, column=1).border = _thin_border()

    ws_segments.cell(row=row, column=2).value = f"=SUM(B5:B9)"
    ws_segments.cell(row=row, column=2).number_format = "#,##0"
    ws_segments.cell(row=row, column=2).font = Font(bold=True)
    ws_segments.cell(row=row, column=2).fill = PatternFill(start_color="00E8F4F0", end_color="00E8F4F0", fill_type="solid")
    ws_segments.cell(row=row, column=2).border = _thin_border()

    ws_segments.cell(row=row, column=5).value = f"=SUM(E5:E9)"
    ws_segments.cell(row=row, column=5).number_format = "#,##0"
    ws_segments.cell(row=row, column=5).font = Font(bold=True)
    ws_segments.cell(row=row, column=5).fill = PatternFill(start_color="00E8F4F0", end_color="00E8F4F0", fill_type="solid")
    ws_segments.cell(row=row, column=5).border = _thin_border()

    # ===== FUNNEL SHEET =====
    ws_funnel = wb.create_sheet("Funnel", 1)
    ws_funnel.column_dimensions["A"].width = 28
    ws_funnel.column_dimensions["B"].width = 16
    ws_funnel.column_dimensions["C"].width = 16
    ws_funnel.column_dimensions["D"].width = 16

    # Title row
    ws_funnel.merge_cells("A1:D1")
    ws_funnel["A1"] = "Sales Funnel Waterfall"
    ws_funnel["A1"].fill = _header_fill()
    ws_funnel["A1"].font = _header_font()
    ws_funnel.row_dimensions[1].height = 20

    # Instruction row
    ws_funnel.merge_cells("A2:D2")
    ws_funnel["A2"] = "Input starting opportunity count and conversion rates. Each stage computed from prior stage."
    ws_funnel["A2"].font = _instruction_font()
    ws_funnel.row_dimensions[2].height = 18

    row = 4
    ws_funnel[f"A{row}"] = "FUNNEL STAGE"
    ws_funnel[f"A{row}"].font = _header_font()
    ws_funnel[f"A{row}"].fill = _header_fill()

    ws_funnel[f"B{row}"] = "Count"
    ws_funnel[f"B{row}"].font = _header_font()
    ws_funnel[f"B{row}"].fill = _header_fill()
    ws_funnel[f"B{row}"].alignment = _center_align()
    ws_funnel[f"B{row}"].border = _thin_border()

    ws_funnel[f"C{row}"] = "Conversion Rate %"
    ws_funnel[f"C{row}"].font = _header_font()
    ws_funnel[f"C{row}"].fill = _header_fill()
    ws_funnel[f"C{row}"].alignment = _center_align()
    ws_funnel[f"C{row}"].border = _thin_border()

    ws_funnel[f"D{row}"] = "ARR Impact (EUR)"
    ws_funnel[f"D{row}"].font = _header_font()
    ws_funnel[f"D{row}"].fill = _header_fill()
    ws_funnel[f"D{row}"].alignment = _center_align()
    ws_funnel[f"D{row}"].border = _thin_border()

    funnel_data = [
        ["Leads Generated", 1000, None, None],
        ["Marketing Qualified Leads", None, 0.25, None],
        ["Sales Qualified Opportunities", None, 0.60, None],
        ["Proposals Sent", None, 0.40, None],
        ["Closed Won", None, 0.50, None],
    ]

    row += 1
    for idx, data_row in enumerate(funnel_data):
        ws_funnel.cell(row=row, column=1).value = data_row[0]
        ws_funnel.cell(row=row, column=1).alignment = _left_align()
        ws_funnel.cell(row=row, column=1).border = _thin_border()

        if idx == 0:
            # Leads: manual input
            ws_funnel.cell(row=row, column=2).value = data_row[1]
            ws_funnel.cell(row=row, column=2).fill = _light_yellow_fill()
            ws_funnel.cell(row=row, column=2).number_format = "0"
            ws_funnel.cell(row=row, column=2).alignment = _right_align()
            ws_funnel.cell(row=row, column=2).border = _thin_border()
        else:
            # Subsequent stages: prior count * conversion rate
            ws_funnel.cell(row=row, column=2).value = f"=B{row - 1}*C{row}"
            ws_funnel.cell(row=row, column=2).number_format = "0"
            ws_funnel.cell(row=row, column=2).alignment = _right_align()
            ws_funnel.cell(row=row, column=2).border = _thin_border()

        if data_row[2] is not None:
            ws_funnel.cell(row=row, column=3).value = data_row[2]
            ws_funnel.cell(row=row, column=3).fill = _light_yellow_fill()
            ws_funnel.cell(row=row, column=3).number_format = "0%"
            ws_funnel.cell(row=row, column=3).alignment = _right_align()
            ws_funnel.cell(row=row, column=3).border = _thin_border()

        ws_funnel.cell(row=row, column=4).value = ""  # ARR impact optional per stage
        ws_funnel.cell(row=row, column=4).number_format = "#,##0"
        ws_funnel.cell(row=row, column=4).alignment = _right_align()
        ws_funnel.cell(row=row, column=4).border = _thin_border()

        row += 1

    # ===== RECONCILIATION SHEET =====
    ws_recon = wb.create_sheet("Reconciliation", 2)
    ws_recon.column_dimensions["A"].width = 35
    ws_recon.column_dimensions["B"].width = 18
    ws_recon.column_dimensions["C"].width = 16

    # Title row
    ws_recon.merge_cells("A1:C1")
    ws_recon["A1"] = "Bottoms-Up vs Top-Down Reconciliation"
    ws_recon["A1"].fill = _header_fill()
    ws_recon["A1"].font = _header_font()
    ws_recon.row_dimensions[1].height = 20

    # Instruction row
    ws_recon.merge_cells("A2:C2")
    ws_recon["A2"] = "Compare bottoms-up segment forecast against top-down target. Input finance stretch. Plan of record owns gap."
    ws_recon["A2"].font = _instruction_font()
    ws_recon.row_dimensions[2].height = 18

    row = 4
    ws_recon[f"A{row}"] = "FORECAST COMPARISON"
    ws_recon[f"A{row}"].font = _header_font()
    ws_recon[f"A{row}"].fill = _header_fill()
    row += 1

    comparison_headers = ["Forecast Source", "Amount (EUR)", "% of Bottoms-Up"]
    for col_idx, header in enumerate(comparison_headers, 1):
        cell = ws_recon.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = _center_align()
        cell.border = _thin_border()

    row += 1
    ws_recon.cell(row=row, column=1).value = "Bottoms-Up (Segments)"
    ws_recon.cell(row=row, column=1).alignment = _left_align()
    ws_recon.cell(row=row, column=1).border = _thin_border()
    ws_recon.cell(row=row, column=2).value = "=Segments!E10"
    ws_recon.cell(row=row, column=2).number_format = "#,##0"
    ws_recon.cell(row=row, column=2).alignment = _right_align()
    ws_recon.cell(row=row, column=2).border = _thin_border()
    ws_recon.cell(row=row, column=3).value = "100%"
    ws_recon.cell(row=row, column=3).number_format = "0%"
    ws_recon.cell(row=row, column=3).alignment = _right_align()
    ws_recon.cell(row=row, column=3).border = _thin_border()
    bottoms_up_row = row
    row += 1

    ws_recon.cell(row=row, column=1).value = "Finance Stretch Target"
    ws_recon.cell(row=row, column=1).alignment = _left_align()
    ws_recon.cell(row=row, column=1).border = _thin_border()
    ws_recon.cell(row=row, column=2).fill = _light_yellow_fill()
    ws_recon.cell(row=row, column=2).number_format = "#,##0"
    ws_recon.cell(row=row, column=2).alignment = _right_align()
    ws_recon.cell(row=row, column=2).border = _thin_border()
    ws_recon.cell(row=row, column=3).value = f"=B{row}/B{bottoms_up_row}"
    ws_recon.cell(row=row, column=3).number_format = "0%"
    ws_recon.cell(row=row, column=3).alignment = _right_align()
    ws_recon.cell(row=row, column=3).border = _thin_border()
    stretch_row = row
    row += 1

    ws_recon.cell(row=row, column=1).value = "Plan of Record"
    ws_recon.cell(row=row, column=1).alignment = _left_align()
    ws_recon.cell(row=row, column=1).border = _thin_border()
    ws_recon.cell(row=row, column=2).fill = _light_yellow_fill()
    ws_recon.cell(row=row, column=2).number_format = "#,##0"
    ws_recon.cell(row=row, column=2).alignment = _right_align()
    ws_recon.cell(row=row, column=2).border = _thin_border()
    ws_recon.cell(row=row, column=3).value = f"=B{row}/B{bottoms_up_row}"
    ws_recon.cell(row=row, column=3).number_format = "0%"
    ws_recon.cell(row=row, column=3).alignment = _right_align()
    ws_recon.cell(row=row, column=3).border = _thin_border()
    plan_row = row
    row += 2

    ws_recon[f"A{row}"] = "PLAN VERSION RECORD"
    ws_recon[f"A{row}"].font = _header_font()
    ws_recon[f"A{row}"].fill = _header_fill()
    row += 1

    version_headers = ["Plan Version", "Amount (EUR)", "Date Created", "Owner"]
    for col_idx, header in enumerate(version_headers, 1):
        cell = ws_recon.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = _center_align()
        cell.border = _thin_border()

    version_data = [
        ["Bottoms-Up Original", f"=B{bottoms_up_row}", None, None],
        ["Finance Stretch", f"=B{stretch_row}", None, None],
        ["Plan of Record", f"=B{plan_row}", None, None],
    ]

    row += 1
    for label, formula, date_val, owner_val in version_data:
        ws_recon.cell(row=row, column=1).value = label
        ws_recon.cell(row=row, column=1).alignment = _left_align()
        ws_recon.cell(row=row, column=1).border = _thin_border()

        ws_recon.cell(row=row, column=2).value = formula
        ws_recon.cell(row=row, column=2).number_format = "#,##0"
        ws_recon.cell(row=row, column=2).alignment = _right_align()
        ws_recon.cell(row=row, column=2).border = _thin_border()

        if label == "Plan of Record":
            ws_recon.cell(row=row, column=3).fill = _light_yellow_fill()
        ws_recon.cell(row=row, column=3).number_format = "YYYY-MM-DD"
        ws_recon.cell(row=row, column=3).alignment = _center_align()
        ws_recon.cell(row=row, column=3).border = _thin_border()

        if label == "Plan of Record":
            ws_recon.cell(row=row, column=4).fill = _light_yellow_fill()
        ws_recon.cell(row=row, column=4).alignment = _left_align()
        ws_recon.cell(row=row, column=4).border = _thin_border()

        row += 1

    row += 2
    ws_recon[f"A{row}"] = "NAMED SCENARIOS"
    ws_recon[f"A{row}"].font = _header_font()
    ws_recon[f"A{row}"].fill = _header_fill()
    row += 1

    scenario_headers = ["Scenario Name", "Owner", "Amount (EUR)", "Confidence Level"]
    for col_idx, header in enumerate(scenario_headers, 1):
        cell = ws_recon.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = _center_align()
        cell.border = _thin_border()

    row += 1
    for _ in range(3):  # Three empty scenario rows for user input
        for col in range(1, 5):
            cell = ws_recon.cell(row=row, column=col)
            if col > 1:
                cell.fill = _light_yellow_fill()
            cell.alignment = _left_align() if col in [1, 2] else _right_align()
            if col == 3:
                cell.number_format = "#,##0"
            cell.border = _thin_border()
        row += 1

    wb.save(str(out_path))


def main():
    """Generate both workbooks into the assets directory."""
    skill_root = Path(__file__).parent.parent
    assets_dir = skill_root / "assets"
    assets_dir.mkdir(exist_ok=True)

    capacity_path = assets_dir / "capacity-model-calculator.xlsx"
    bottoms_up_path = assets_dir / "bottoms-up-planning-sheet.xlsx"

    print(f"Generating capacity-model-calculator.xlsx...")
    generate_capacity_model_calculator(capacity_path)
    print(f"  Created: {capacity_path}")

    print(f"Generating bottoms-up-planning-sheet.xlsx...")
    generate_bottoms_up_planning_sheet(bottoms_up_path)
    print(f"  Created: {bottoms_up_path}")

    print("Done.")


if __name__ == "__main__":
    main()
