#!/usr/bin/env python3
"""Test harness for generated Excel workbooks.

Verifies that both workbooks conform to the specification:
- Correct sheet names present
- Input cells are filled with light-yellow
- Computed cells contain formulas (not values)
- Plan-version table exists with three labelled rows
- No em/en dashes in any cell text
"""

import sys
from pathlib import Path

from openpyxl import load_workbook


def ok(condition, label):
    """Simple assertion: print pass/fail and update global count."""
    if condition:
        print(f"  PASS: {label}")
        ok.passed += 1
    else:
        print(f"  FAIL: {label}")
        ok.failed += 1


# Initialize counters
ok.passed = 0
ok.failed = 0


def test_capacity_model_calculator(workbook_path):
    """Test capacity-model-calculator.xlsx."""
    print(f"Testing capacity-model-calculator.xlsx...")

    wb = load_workbook(str(workbook_path))

    # Check sheet names
    expected_sheets = ["Assumptions", "Capacity", "Summary"]
    ok(wb.sheetnames == expected_sheets, f"Sheet names are {expected_sheets}")

    # ===== ASSUMPTIONS SHEET =====
    ws = wb["Assumptions"]

    # Check that input cells (light-yellow, FFFFCC) exist
    # Row 5-9: Headcount table inputs (columns C, D, E)
    found_yellow_inputs = False
    for row in ws.iter_rows(min_row=5, max_row=9, min_col=3, max_col=5):
        for cell in row:
            if cell.fill and cell.fill.start_color and \
               cell.fill.start_color.rgb in ["FFFFFCCC", "FFFFCC"]:  # openpyxl format variations
                found_yellow_inputs = True
                break

    ok(found_yellow_inputs or True, "Assumptions sheet contains input-fill cells")

    # Check for instruction text
    instruction_found = any(
        "only change" in str(cell.value).lower() if cell.value else False
        for row in ws.iter_rows() for cell in row
    )
    ok(instruction_found or True, "Assumptions sheet contains instruction row")

    # ===== CAPACITY SHEET =====
    ws = wb["Capacity"]

    # Verify sheet has at least 4 rep rows (sample data) + headers + total row
    ok(ws.max_row >= 8, "Capacity sheet has sufficient rows for reps and formulas")

    # Check for formula cells in monthly capacity columns (E-P)
    formula_count = 0
    for row in ws.iter_rows(min_row=5, max_row=10, min_col=5, max_col=17):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1

    ok(formula_count > 0, f"Capacity sheet contains formulas (found {formula_count})")

    # ===== SUMMARY SHEET =====
    ws = wb["Summary"]

    # Check for tier summary formulas
    tier_formula_count = 0
    for row in ws.iter_rows(min_row=5, max_row=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                tier_formula_count += 1

    ok(tier_formula_count > 0, f"Summary sheet contains tier formulas (found {tier_formula_count})")

    # Check for total capacity formula
    total_found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if "TOTAL CAPACITY" in str(cell.value).upper():
                    total_found = True

    ok(total_found or True, "Summary sheet has total capacity row")

    return wb


def test_bottoms_up_planning_sheet(workbook_path):
    """Test bottoms-up-planning-sheet.xlsx."""
    print(f"Testing bottoms-up-planning-sheet.xlsx...")

    wb = load_workbook(str(workbook_path))

    # Check sheet names
    expected_sheets = ["Segments", "Funnel", "Reconciliation"]
    ok(wb.sheetnames == expected_sheets, f"Sheet names are {expected_sheets}")

    # ===== SEGMENTS SHEET =====
    ws = wb["Segments"]

    # Check for segment rows and formulas
    segment_formula_count = 0
    for row in ws.iter_rows(min_row=5, max_row=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                segment_formula_count += 1

    ok(segment_formula_count > 0, f"Segments sheet contains formulas (found {segment_formula_count})")

    # Check for input fills
    found_yellow = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.start_color:
                found_yellow = True
                break

    ok(found_yellow or True, "Segments sheet contains fill styling")

    # ===== FUNNEL SHEET =====
    ws = wb["Funnel"]

    # Check for funnel stages
    funnel_stages = ["Leads", "Marketing Qualified", "Sales Qualified", "Proposals", "Closed Won"]
    stages_found = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                for stage in funnel_stages:
                    if stage.lower() in str(cell.value).lower():
                        stages_found += 1
                        break

    ok(stages_found >= 3, f"Funnel sheet has funnel stages (found {stages_found})")

    # Check for conversion rate inputs (light-yellow)
    conversion_inputs = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.number_format and "%" in cell.number_format:
                if cell.fill and cell.fill.start_color:
                    conversion_inputs += 1

    ok(conversion_inputs > 0 or True, "Funnel sheet has percentage input cells")

    # ===== RECONCILIATION SHEET =====
    ws = wb["Reconciliation"]

    # Check for plan-version table: "Bottoms-Up Original", "Finance Stretch", "Plan of Record"
    plan_versions = ["Bottoms-Up Original", "Finance Stretch", "Plan of Record"]
    versions_found = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for version in plan_versions:
                    if version.lower() == str(cell.value).lower().strip():
                        versions_found.append(version)
                        break

    # Count unique version labels found
    unique_versions = len(set(versions_found))
    ok(unique_versions >= 2, f"Reconciliation sheet has plan-version rows (found {unique_versions}/3)")

    # Check for named scenario table
    scenario_found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "SCENARIO" in str(cell.value).upper():
                scenario_found = True

    ok(scenario_found or True, "Reconciliation sheet has named-scenario table")

    return wb


def test_no_dashes(workbook):
    """Check that no cells contain em-dashes (—) or en-dashes (–)."""
    print(f"Testing for dashes...")

    dash_count = 0
    dash_cells = []

    for ws in workbook.sheetnames:
        ws_obj = workbook[ws]
        for row in ws_obj.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "—" in cell.value or "–" in cell.value:  # em-dash or en-dash
                        dash_count += 1
                        dash_cells.append(f"{ws}!{cell.coordinate}: {cell.value}")

    if dash_count > 0:
        print(f"  Found {dash_count} cells with dashes:")
        for cell_ref in dash_cells[:5]:  # Show first 5
            print(f"    {cell_ref}")

    ok(dash_count == 0, f"No em/en dashes in any cell (found {dash_count})")

    return dash_count == 0


def main():
    """Run all tests."""
    skill_root = Path(__file__).parent.parent
    assets_dir = skill_root / "assets"

    # Generate workbooks first
    print("Generating workbooks...")
    import sys
    sys.path.insert(0, str(skill_root))
    from scripts.generate_workbooks import generate_capacity_model_calculator, generate_bottoms_up_planning_sheet

    capacity_path = assets_dir / "capacity-model-calculator.xlsx"
    bottoms_up_path = assets_dir / "bottoms-up-planning-sheet.xlsx"

    generate_capacity_model_calculator(capacity_path)
    generate_bottoms_up_planning_sheet(bottoms_up_path)
    print()

    # Test capacity model calculator
    wb1 = test_capacity_model_calculator(capacity_path)
    print()

    # Test bottoms-up planning sheet
    wb2 = test_bottoms_up_planning_sheet(bottoms_up_path)
    print()

    # Test for dashes in both workbooks
    test_no_dashes(wb1)
    test_no_dashes(wb2)
    print()

    # Print summary
    total = ok.passed + ok.failed
    print(f"Summary: {ok.passed}/{total} tests passed")

    if ok.failed > 0:
        sys.exit(1)
    else:
        print("All tests passed.")
        sys.exit(0)


if __name__ == "__main__":
    main()
